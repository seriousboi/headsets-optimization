from openpyxl import load_workbook



def getData():
    data = load_workbook("data.xlsx")

    durationsSheet = data["DUREES"]
    capacitiesSheet = data["CAPACITES"]
    demandsSheet = data["DEMANDES"]
    distancesSheet = data["DISTANCES"]

    headsetsAmount = 3
    factoriesAmount = 3
    citiesAmount = 21

    durations = []
    for factory in range(factoriesAmount):
        durations += [[]]
        for headset in range(headsetsAmount):
            durations[factory] += [durationsSheet.cell(factory+3,headset+2).value]

    capacities = []
    for factory in range(factoriesAmount):
        capacities += [capacitiesSheet.cell(factory+2,2).value]

    demands = []
    for city in range(citiesAmount):
        demands += [[]]
        for headset in range(headsetsAmount):
            demands[city] += [demandsSheet.cell(city+2,headset+3).value]

    distances = []
    for city in range(citiesAmount):
        distances += [[]]
        for factory in range(factoriesAmount):
            distances[city] += [distancesSheet.cell(city+2,factory+2).value]

    travelCost = distancesSheet.cell(2,6).value

    return durations , capacities , demands , distances , travelCost
